'''
author == Wenchen Chen
encoding  = utf-8
'''
'''
修改记录
更改了网络重建的过程
更改曲面的孔径大小
对于给定的结构参数，优化时，不能更改结构参数
'''

import tensorflow as tf
import numpy as np
import pandas as pd
import time
from openpyxl import load_workbook
from classcodev import codev
import shutil
import os

def readxlsx(filename, sheetname):  #  定义读取文件的函数，按照行读取数据
    wb = load_workbook(filename)
    ws = wb[sheetname]
    columns = ws.max_column
    rows = ws.max_row
    Data = np.zeros([rows, columns], dtype='float32')
    # rowdata = []
    for i in range(1, rows + 1):
        for j in range(1, columns + 1):
            cellvalue = ws.cell(row=i, column=j).value
            Data[i - 1, j - 1] = cellvalue
    return Data

def mycopyfile(srcfile, dstpath):  # 复制函数
    if not os.path.isfile(srcfile):
        print("%s not exist!" % (srcfile))
    else:
        fpath, fname = os.path.split(srcfile)  # 分离文件名和路径
        if not os.path.exists(dstpath):
            os.makedirs(dstpath)  # 创建路径
        shutil.copy(srcfile, dstpath + fname)  # 复制文件
        print("copy %s -> %s" % (srcfile, dstpath + fname))

# 按照列进行归一化，python中axis=0代表列，axis=1代表行
def premnmx_col(data):
    col_max = data.max(axis=0)
    col_min = data.min(axis=0)
    a = (data - 0.5 * (col_min + col_max)) / (0.5 * (col_max - col_min))
    return col_min, col_max, a


def tramnmx_col(data, min, max):
    a = (data - 0.5 * (min + max)) / (0.5 * (max - min))
    return a


def postmnmx_col(data, min, max):
    a = data * 0.5 * (max - min) + 0.5 * (min + max)
    return a

def data_left_right(data_min,data_max):
    data_left = data_min + 1/8*(data_max-data_min)
    data_right = data_max - 1/8*(data_max-data_min)
    return data_left, data_right


def extreme_data_generate(data_min,data_max):
    data_left = data_min + 1/8*(data_max-data_min)
    data_right = data_max - 1/8*(data_max-data_min)
    sigma = 0.05*(yde_s2_max-yde_s2_min)
    model = np.random.random()
    if model <= 0.5:
        extreme_data = np.random.normal(data_left,sigma)
    else:
        extreme_data = np.random.normal(data_right, sigma)
    if extreme_data <= data_min:
        extreme_data = data_min
    if extreme_data >= data_max:
        extreme_data = data_max
    return extreme_data
'''
模块1：生成随机的系统参数，自由度越小，搜索空间越小，最小的输入为3个，最大的自由度为6.
'''
work_model = 'unopt_model'
output_num = 50
total_predict_sys = 10000



'''
生成随机系统参数
'''
x_fov_min = 1
x_fov_max = 5
y_fov_min = 1
y_fov_max = 5
epd_min = 15
epd_max = 25

yde_s2_min = 50
yde_s2_max = 100
zde_s2_min = 150
zde_s2_max = 250
yde_s4_min = -50
yde_s4_max = 0
zde_s4_min = 150
zde_s4_max = 250
yde_si_min = -60
yde_si_max = -16
zde_si_min = -30
zde_si_max = 30
max_mtf = 5

struct_model = np.array(np.zeros([1,6])).reshape(1,6)
if yde_s2_min == yde_s2_max:
    struct_model[0,0] = 1
if zde_s2_min == zde_s2_max:
    struct_model[0,1] = 2
if yde_s4_min == yde_s4_max:
    struct_model[0,2] = 3
if zde_s4_min == zde_s4_max:
    struct_model[0,3] = 4
if yde_si_min == yde_si_max:
    struct_model[0,4] = 5
if zde_si_min == zde_si_max:
    struct_model[0,5] = 6

freeze_data = {1:'ydc s2 100',2:'zdc s2 100',3:'ydc s4 100',4:'zdc s4 100',5:'ydc si 100',6:'zdc si 100'}



unopt_sys_file = r'D:\cwc\three_mirror_starting_point_data_set\train_models\three_mirror_strarting_point_prediction\unoptimized_system'
opt_sys_file = r'D:\cwc\three_mirror_starting_point_data_set\train_models\three_mirror_strarting_point_prediction\optimized_system'
output_file_name_ini = unopt_sys_file

# 生成输入参数
DNN_input = np.zeros([total_predict_sys,9])
np.random.seed(3)


# 极端情况下的输入


for i in range(total_predict_sys):
    DNN_input[i, 0] = np.round(x_fov_min + (x_fov_max - x_fov_min) * np.random.rand(1), 15)
    DNN_input[i, 1] = np.round(y_fov_min + (y_fov_max - y_fov_min) * np.random.rand(1), 15)
    DNN_input[i, 2] = np.round(epd_min + (epd_max - epd_min) * np.random.rand(1), 15)
    DNN_input[i, 3] = extreme_data_generate(yde_s2_min,yde_s2_max)
    DNN_input[i, 4] = extreme_data_generate(zde_s2_min,zde_s2_max)
    DNN_input[i, 5] = extreme_data_generate(yde_s4_min,yde_s4_max)
    DNN_input[i, 6] = extreme_data_generate(zde_s4_min,zde_s4_max)
    DNN_input[i, 7] = extreme_data_generate(yde_si_min,yde_si_max)
    DNN_input[i, 8] = extreme_data_generate(zde_si_min,zde_si_max)

if work_model == 'unopt_model':
    dnn_random_input_file = unopt_sys_file+'\predict_sys_data.xlsx'
else:
    dnn_random_input_file = opt_sys_file+'\predict_sys_data.xlsx'
data= pd.DataFrame(DNN_input)
writer = pd.ExcelWriter(dnn_random_input_file)
data.to_excel(writer,'Sheet1',float_format='%.20f')
writer.save()
writer.close()

'''
模块2：网络再现
'''

'''
打开训练集文件
'''
dataset_filename = r'D:\cwc\three_mirror_multi_specification_code\train_dataset_total.xlsx'
file_sheet = 'Sheet2'
total_data = readxlsx(dataset_filename, file_sheet)
input_data_original = total_data[:,0:9]
output_data_original = total_data[:,9:]
min_input, max_input, x_vals = premnmx_col(input_data_original)
min_output, max_output, y_vals = premnmx_col(output_data_original)
y_vals = np.nan_to_num(y_vals)

np.random.seed(2)
train_indices = np.random.choice(len(x_vals), 90000, replace=False)
test_indices = np.array(list(set(range(len(x_vals))) - set(train_indices)))

x_vals_train = x_vals[train_indices]
x_vals_test = x_vals[test_indices]
y_vals_train = y_vals[train_indices]
y_vals_test = y_vals[test_indices]

train_data = total_data[train_indices]
test_data = total_data[test_indices]

total_train_num = len(x_vals_train)

predict_system_data = DNN_input

'''
打开code v准备相关工作
'''
test_time_start = time.time()
test_system_performance_info = np.zeros([total_predict_sys, 21])

sur1_num = 2
sur2_num = 3
sur3_num = 4
sur4_num = 5



initial_lens_filename = r'D:\cwc\three_mirror_starting_point_data_set\test_lens\test.len'

cv = codev()
cv.cvon()
cv.cvopen(initial_lens_filename)

'''
模型复现
'''
input_node = 9
hidden1_node = 40
hidden2_node = 60
hidden3_node = 80
hidden4_node = 100
hidden5_node = 120
hidden6_node = 140
hidden7_node = 160
hidden8_node = 180
hidden9_node = 200
hidden10_node = 240
hidden11_node = 260
hidden12_node = 280
hidden13_node = 300
hidden14_node = 300
hidden15_node = 300
hidden16_node = 300
hidden17_node = 300
hidden18_node = 300
hidden19_node = 300
hidden20_node = 300
hidden21_node = 300
hidden22_node = 300
hidden23_node = 280
hidden24_node = 260
hidden25_node = 240
hidden26_node = 220
hidden27_node = 200
hidden28_node = 180
hidden29_node = 160
hidden30_node = 150
hidden31_node = 140
hidden32_node = 130
hidden33_node = 120
hidden34_node = 110
hidden35_node = 100
hidden36_node = 80
hidden37_node = 70
hidden38_node = 60
hidden39_node = 55
hidden40_node = 52
output_node = 52

tf_x = tf.placeholder(shape=[None, 9], dtype=tf.float32)  # input x
l1 = tf.layers.dense(tf_x, hidden1_node, tf.nn.tanh)
l2 = tf.layers.dense(l1, hidden2_node, tf.nn.tanh)
l3 = tf.layers.dense(l2, hidden3_node, tf.nn.tanh)
l4 = tf.layers.dense(l3, hidden4_node, tf.nn.tanh)
l5 = tf.layers.dense(l4, hidden5_node, tf.nn.tanh)
l6 = tf.layers.dense(l5, hidden6_node, tf.nn.tanh)
l7 = tf.layers.dense(l6, hidden7_node, tf.nn.tanh)
l8 = tf.layers.dense(l7, hidden8_node, tf.nn.tanh)
l9 = tf.layers.dense(l8, hidden9_node, tf.nn.tanh)
l10 = tf.layers.dense(l9, hidden10_node, tf.nn.tanh)
l11 = tf.layers.dense(l10, hidden11_node, tf.nn.tanh)
l12 = tf.layers.dense(l11, hidden12_node, tf.nn.tanh)
l13 = tf.layers.dense(l12, hidden13_node, tf.nn.tanh)
l14 = tf.layers.dense(l13, hidden14_node, tf.nn.tanh)
l15 = tf.layers.dense(l14, hidden15_node, tf.nn.tanh)
l16 = tf.layers.dense(l15, hidden16_node, tf.nn.tanh)
l17 = tf.layers.dense(l16, hidden17_node, tf.nn.tanh)
l18 = tf.layers.dense(l17, hidden18_node, tf.nn.tanh)
l19 = tf.layers.dense(l18, hidden19_node, tf.nn.tanh)
l20 = tf.layers.dense(l19, hidden20_node, tf.nn.tanh)
l21 = tf.layers.dense(l20, hidden21_node, tf.nn.tanh)
l22 = tf.layers.dense(l21, hidden22_node, tf.nn.tanh)
l23 = tf.layers.dense(l22, hidden23_node, tf.nn.tanh)
l24 = tf.layers.dense(l23, hidden24_node, tf.nn.tanh)
l25 = tf.layers.dense(l24, hidden25_node, tf.nn.tanh)
l26 = tf.layers.dense(l25, hidden26_node, tf.nn.tanh)
l27 = tf.layers.dense(l26, hidden27_node, tf.nn.tanh)
l28 = tf.layers.dense(l27, hidden28_node, tf.nn.tanh)
l29 = tf.layers.dense(l28, hidden29_node, tf.nn.tanh)
l30 = tf.layers.dense(l29, hidden30_node, tf.nn.tanh)
l31 = tf.layers.dense(l30, hidden31_node, tf.nn.tanh)
l32 = tf.layers.dense(l31, hidden32_node, tf.nn.tanh)
l33 = tf.layers.dense(l32, hidden33_node, tf.nn.tanh)
l34 = tf.layers.dense(l33, hidden34_node, tf.nn.tanh)
l35 = tf.layers.dense(l34, hidden35_node, tf.nn.tanh)
l36 = tf.layers.dense(l35, hidden36_node, tf.nn.tanh)
l37 = tf.layers.dense(l36, hidden37_node, tf.nn.tanh)
l38 = tf.layers.dense(l37, hidden38_node, tf.nn.tanh)
l39 = tf.layers.dense(l38, hidden39_node, tf.nn.tanh)
l40 = tf.layers.dense(l39, hidden40_node, tf.nn.tanh)
actual_output = tf.layers.dense(l40, output_node)
sess = tf.Session()
saver = tf.train.Saver()
saver.restore(sess, 'starting_point_model50_6/model50_6_params')

problematic_sys_count = 0
obstructed_sys_count = 0
good_system_count = 0
predict_start_time = time.time()
for i in range(total_predict_sys):  # i的取值是从0开始的
    evalx = np.array(predict_system_data[i, :]).reshape(1, 9)
    evalx_ = tramnmx_col(evalx, min_input,max_input)
    pred_ = sess.run(actual_output, {tf_x: evalx_})

    pred = postmnmx_col(pred_, min_output, max_output)
    cv.cvcmd('in CV_MACRO:cvsetfield Y ' + str(evalx[0, 1]) + ' F2')
    cv.cvcmd('in CV_MACRO:cvsetfield Y ' + str(-evalx[0, 1]) + ' F3')
    cv.cvcmd('in CV_MACRO:cvsetfield Y ' + str(evalx[0, 1]) + ' F5')
    cv.cvcmd('in CV_MACRO:cvsetfield Y ' + str(-evalx[0, 1]) + ' F6')

    cv.cvcmd('in CV_MACRO:cvsetfield X ' + str(evalx[0, 0]) + ' F4')
    cv.cvcmd('in CV_MACRO:cvsetfield X ' + str(evalx[0, 0]) + ' F5')
    cv.cvcmd('in CV_MACRO:cvsetfield X ' + str(evalx[0, 0]) + ' F6')

    cv.cvcmd('EPD '+str(evalx[0, 2]))
    # 输入结构参数

    cv.cvcmd('YDE S' + str(sur1_num) + ' ' + str(pred[0, 0]))
    cv.cvcmd('ZDE S' + str(sur1_num) + ' ' + str(pred[0, 1]))
    cv.cvcmd('ADE S' + str(sur1_num) + ' ' + str(pred[0, 2]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C4 ' + str(pred[0, 3]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C6 ' + str(pred[0, 4]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C8 ' + str(pred[0, 5]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C10 ' + str(pred[0, 6]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C11 ' + str(pred[0, 7]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C13 ' + str(pred[0, 8]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C15 ' + str(pred[0, 9]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C17 ' + str(pred[0, 10]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C19 ' + str(pred[0, 11]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C21 ' + str(pred[0, 12]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C22 ' + str(pred[0, 13]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C24 ' + str(pred[0, 14]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C26 ' + str(pred[0, 15]))
    cv.cvcmd('SCO S' + str(sur1_num) + ' C28 ' + str(pred[0, 16]))

    cv.cvcmd('ADE S' + str(sur2_num) + ' ' + str(pred[0, 17]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C4 ' + str(pred[0, 18]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C6 ' + str(pred[0, 19]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C8 ' + str(pred[0, 20]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C10 ' + str(pred[0, 21]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C11 ' + str(pred[0, 22]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C13 ' + str(pred[0, 23]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C15 ' + str(pred[0, 24]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C17 ' + str(pred[0, 25]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C19 ' + str(pred[0, 26]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C21 ' + str(pred[0, 27]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C22 ' + str(pred[0, 28]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C24 ' + str(pred[0, 29]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C26 ' + str(pred[0, 30]))
    cv.cvcmd('SCO S' + str(sur2_num) + ' C28 ' + str(pred[0, 31]))

    cv.cvcmd('YDE S' + str(sur3_num) + ' ' + str(pred[0, 32]))
    cv.cvcmd('ZDE S' + str(sur3_num) + ' ' + str(pred[0, 33]))
    cv.cvcmd('ADE S' + str(sur3_num) + ' ' + str(pred[0, 34]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C4 ' + str(pred[0, 35]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C6 ' + str(pred[0, 36]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C8 ' + str(pred[0, 37]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C10 ' + str(pred[0, 38]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C11 ' + str(pred[0, 39]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C13 ' + str(pred[0, 40]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C15 ' + str(pred[0, 41]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C17 ' + str(pred[0, 42]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C19 ' + str(pred[0, 43]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C21 ' + str(pred[0, 44]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C22 ' + str(pred[0, 45]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C24 ' + str(pred[0, 46]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C26 ' + str(pred[0, 47]))
    cv.cvcmd('SCO S' + str(sur3_num) + ' C28 ' + str(pred[0, 48]))

    cv.cvcmd('YDE S' + str(sur4_num) + ' ' + str(pred[0, 49]))
    cv.cvcmd('ZDE S' + str(sur4_num) + ' ' + str(pred[0, 50]))
    cv.cvcmd('ADE S' + str(sur4_num) + ' ' + str(pred[0, 51]))

    cv.cvcmd('REX S1 200')
    cv.cvcmd('REY S1 200')
    cv.cvcmd('REX S2 200')
    cv.cvcmd('REY S2 200')
    cv.cvcmd('REX S4 200')
    cv.cvcmd('REY S4 200')
    cv.cvcmd('REX S5 200')
    cv.cvcmd('REY S5 200')

    cv.cvcmd('YDC S2  100')
    cv.cvcmd('zDC S2  100')
    cv.cvcmd('YDC s4  100')
    cv.cvcmd('zDC s4  100')
    cv.cvcmd('YDC S5  100')
    cv.cvcmd('zDC S5  100')

    #如果对结构参数有要求，将预测出来的系统的结构参数进行替换
    # if struct_model.all() ==False:
    #     print('no specific structure data is needed ')
    # else:
    #
    #     cv.cvcmd('YDE S' + str(sur1_num) + ' ' + str(evalx[0, 3]))
    #     cv.cvcmd('ZDE S' + str(sur1_num) + ' ' + str(evalx[0, 4]))
    #     cv.cvcmd('YDE S' + str(sur3_num) + ' ' + str(evalx[0, 5]))
    #     cv.cvcmd('ZDE S' + str(sur3_num) + ' ' + str(evalx[0, 6]))
    #     cv.cvcmd('YDE S' + str(sur4_num) + ' ' + str(evalx[0, 7]))
    #     cv.cvcmd('ZDE S' + str(sur4_num) + ' ' + str(evalx[0, 8]))


    if work_model == 'unopt_model':
        pass
    else:
        for mm in range(6):
            if struct_model[0, mm] != 0:
                print(freeze_data[mm+1])
                cv.cvcmd(freeze_data[mm + 1])
        cv.cvmacro(r'D:\cwc\codev_2007a\codev_2007a\cvmacro\three_mirror_advanced\three_mirror_starting_point_optimization.seq')

    # 像质评价模块
    # 设置渐晕

    cv.cvcmd('"E:\CODEV\macro\setvig.seq" 1e-007 0.1 100 NO YES ;GO')
    # 相对畸变33333
    cv.cvcmd('buf del ba')
    cv.cvcmd('buf yes')
    cv.cvmacro(r'D:\cwc\codev_2007a\codev_2007a\cvmacro\three_mirror_advanced\three_mirror_relative_distortion_ver2.seq')
    relative_disrtortion = np.array(cv.cvbuf(3, 1, 1, 1, 2)).reshape(1, 2)
    test_system_performance_info[i, 5:7] = relative_disrtortion
    print('the relative distortion of the system is: ', relative_disrtortion)

    # 误差函数  1111
    cv.cvcmd('buf del ba')
    cv.cvcmd('buf yes')
    cv.cvmacro(r'D:\cwc\codev_2007a\codev_2007a\cvmacro\three_mirror_advanced\error_function.seq')
    error_function = np.array(cv.cvbuf(1, 1, 1, 1, 1)).reshape(1, 1)
    test_system_performance_info[i, 1] = error_function
    print('the error_function of the system is: ', error_function)


    # RMS弥散斑大小和 绝对畸变2222

    cv.cvcmd('buf del ba')
    cv.cvcmd('buf yes')
    cv.cvmacro(r'D:\cwc\codev_2007a\codev_2007a\cvmacro\performance_analysis_resulttest_ver2.seq', [100])
    performance = np.array(cv.cvbuf(2, 1, 1, 1, 3)).reshape(1, 3)
    test_system_performance_info[i, 2:5] = performance
    print('the performance of the system is: ', performance)

    # 遮拦判断以及各个距离大小44444
    cv.cvcmd('buf del ba')
    cv.cvcmd('buf yes')
    cv.cvmacro(r'D:\cwc\codev_2007a\codev_2007a\cvmacro\three_mirror_advanced\three_mirror_obstruct_judge.seq')
    judge = np.array(cv.cvbuf(4, 1, 1, 1, 12)).reshape(1, 12)
    test_system_performance_info[i, 7:19] = judge
    print('the judge of the system is: ', judge)

    '''
    计算当前系统体积 555555
    '''
    cv.cvcmd('buf del ba')
    cv.cvcmd('buf yes')
    cv.cvmacro(r'D:\cwc\codev_2007a\codev_2007a\cvmacro\three_mirror_advanced\three_mirror_volume3.seq')
    volume = np.array(cv.cvbuf(5, 1, 1, 1, 1)).reshape(1, 1)
    test_system_performance_info[i,19] = volume
    if volume <= 1e-5:
        cv.cvcmd('CIR S1 300')
        cv.cvcmd('buf del ba')
        cv.cvcmd('buf yes')
        cv.cvmacro(r'D:\cwc\codev_2007a\codev_2007a\cvmacro\three_mirror_advanced\three_mirror_volume3.seq')
        volume = np.array(cv.cvbuf(5, 1, 1, 1, 1)).reshape(1, 1)
        test_system_performance_info[i, 19] = volume


    print('the volume of the system is: ', volume)
    test_system_performance_info[i, 0] = i

    '''
    计算当前系统的mtf
    '''
    cv.cvcmd('buf del ba')
    cv.cvcmd('buf yes')
    cv.cvmacro(r'D:\cwc\codev_2007a\codev_2007a\cvmacro\three_mirror_advanced\get_mtf.seq',[max_mtf])
    mtf_value = np.array(cv.cvbuf(6, 1, 2, 1, 6)).reshape(2, 6)
    min_mtf = np.min(mtf_value)
    test_system_performance_info[i,20] = min_mtf
    print('the min mtf is: ',min_mtf)



    if work_model == 'unopt_model':
        print('the system is unoptimized')
        output_file_name_ini = unopt_sys_file
    else:
        print('the system is optimized')
        output_file_name_ini = opt_sys_file

    if performance[0][0] == 1:
        print('this is a ray tracing error system')
        output_file_name_ini = output_file_name_ini + '/predict_system_problematic_'
        output_file_name = output_file_name_ini + str(i) + '.len'
        cv.cvsave(output_file_name)
        problematic_sys_count = problematic_sys_count + 1
    elif judge[0][0] == 1:
        print('this is a obstructed system')
        obstruct_ini = output_file_name_ini + '/predict_system_obstruct_'
        obstruct = obstruct_ini + str(i) + '.len'
        cv.cvsave(obstruct)
        obstructed_sys_count = obstructed_sys_count + 1
    else:
        print('this is a good system')
        final_lens_filename_ini = output_file_name_ini + r'/predict_system_good_'
        final_lens_filename = final_lens_filename_ini + str(i) + '.len'
        cv.cvsave(final_lens_filename)
        good_system_count = good_system_count + 1

    tf.reset_default_graph()
test_time_end = time.time()

if work_model =='unopt_model':
    test_system_performance_info_file = unopt_sys_file
    test_system_performance_info_file = test_system_performance_info_file + r'\unopt_test_system_performance_info.xlsx'
else:
    test_system_performance_info_file = opt_sys_file
    test_system_performance_info_file = test_system_performance_info_file + r'\opt_test_system_performance_info.xlsx'
data = pd.DataFrame(test_system_performance_info)
writer = pd.ExcelWriter(test_system_performance_info_file)  # 写入Excel文件
data.to_excel(writer, 'Sheet1', float_format='%.15f')  # ‘page_1’是写入excel的sheet名
writer.save()
writer.close()


total_time = round((test_time_end - test_time_start) / 3600, 3)
print('the test time is:', str(total_time), 'hours')  #
print('the problem system number is:',problematic_sys_count)
print('the good system number is:',good_system_count)
print('the obstructed system number is:',obstructed_sys_count)



'''
对提取后的数据进行处理
'''

ss = {1:"error_function",2:'ray_tracing_flag',3:'rms',4:'Absolute_distortion',5:'X_relative_distortion',
      6:'y_relative_distortion',7:'obscuration_flag',8:'s2_yde',9:'s4_yde',10:'s2_s4_yde',11:'s2_zde',
      12:'s4_zde',13:'s2_s4_zde',14:'distance1',15:'distance2',16:'distance3',17:'distance4',18:'distance5',
      19:'volume',20:'mtf'}

''' 首先对数据进行过滤，只留下，光线追迹正常，无遮拦的系统'''
good_performance_data =[]
for z in range(total_predict_sys):
    if test_system_performance_info[z,2] == 0 and test_system_performance_info[z,7] == 0:
        good_performance_data.append(test_system_performance_info[z,:])
good_performance_data = np.array(good_performance_data).reshape(-1,21)


for zz in range(1,21):
    if zz == 20:
        sort = -good_performance_data.T
    else:
        sort = good_performance_data[:,:zz+1].T
    sort_index = np.lexsort(sort)
    sorted_performance_data = good_performance_data[sort_index]
    print('当前参数指标：',ss[zz],'   输出个个数为：',output_num)

    mm = sorted_performance_data[0:output_num+1,zz]
    print(mm)

    '''
    输出系统的名称，并将同一类型系统保存在同一个文件夹中
    '''
    if work_model =='unopt_model':
        new_file_1 = r'D:/cwc/three_mirror_starting_point_data_set/train_models/three_mirror_strarting_point_prediction/unoptimized_system/'
        old_file_1 =  r'D:\cwc\three_mirror_starting_point_data_set\train_models\three_mirror_strarting_point_prediction\unoptimized_system'
    else:
        new_file_1 = r'D:/cwc/three_mirror_starting_point_data_set/train_models/three_mirror_strarting_point_prediction/optimized_system/'
        old_file_1 =  r'D:\cwc\three_mirror_starting_point_data_set\train_models\three_mirror_strarting_point_prediction\optimized_system'
    new_file = new_file_1 + ss[zz] + '/'
    for i in range(output_num):
        old_file = old_file_1 + '\predict_system_good_' + str(int(sorted_performance_data[i, 0])) + '.len'
        mycopyfile(old_file, new_file)
predict_end_time = time.time()
total_predict_time =predict_end_time- predict_start_time

print('total predict time: ',total_predict_time)